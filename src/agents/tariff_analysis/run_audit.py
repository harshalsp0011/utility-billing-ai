import pandas as pd
import logging
import os
from datetime import datetime
from pathlib import Path
from engine import AuditEngine

# Optional: provide CSV override via environment variable AUDIT_CSV_PATH
# or account override via AUDIT_ACCOUNT

# Setup Logging
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

# --- CONFIGURATION ---
# Resolve paths relative to project root for portability.
def _resolve_paths():
    """Resolve input/output paths with fallbacks.

    Priority order for CSV:
    1. AUDIT_CSV_PATH environment variable (absolute or relative)
    2. data/raw/sample_bill.xlsx - Sheet1.csv (legacy exported CSV name)
    3. data/raw/sample_bill.xlsx (we will read first sheet if Excel)
    """
    root = Path(__file__).resolve().parents[3]
    processed_dir = root / "data" / "processed"
    raw_dir = root / "data" / "raw"

    env_csv = os.getenv("AUDIT_CSV_PATH")
    if env_csv:
        csv_candidate = Path(env_csv)
        if not csv_candidate.is_absolute():
            csv_candidate = root / csv_candidate
    else:
        # legacy exported CSV name
        csv_candidate = raw_dir / "sample_bill.xlsx - Sheet1.csv"
        if not csv_candidate.exists():
            # fallback to original Excel file
            excel_fallback = raw_dir / "sample_bill.xlsx"
            csv_candidate = excel_fallback

    logic_path = processed_dir / "tariff_definitions.json"
    output_dir = processed_dir
    return csv_candidate, logic_path, output_dir

CSV_FILE_PATH, LOGIC_FILE_PATH, OUTPUT_DIR = _resolve_paths()
TARGET_ACCOUNT = os.getenv("AUDIT_ACCOUNT", "1120031219")  # Account override via env
FORCED_SC_CODE = os.getenv("AUDIT_SC_OVERRIDE", "SC1")      # Service class override via env

def run_audit_pipeline():
    logger.info(f"🚀 Starting Audit Pipeline")
    logger.info(f"📂 Data Source: {CSV_FILE_PATH}")
    logger.info(f"🎯 Target Account: {TARGET_ACCOUNT}")
    logger.info(f"🔧 Forcing Service Class: {FORCED_SC_CODE}")

    # 1. Initialize Engine
    if not LOGIC_FILE_PATH.exists():
        logger.error(f"❌ Logic file {LOGIC_FILE_PATH} not found. Run extract_logic.py first.")
        return

    engine = AuditEngine(str(LOGIC_FILE_PATH))

    # 2. Load Data (Replacing the old mock function)
    if not CSV_FILE_PATH.exists():
        logger.error(
            f"❌ Input file not found: {CSV_FILE_PATH}\n"
            "Place your CSV/Excel in data/raw or set AUDIT_CSV_PATH."
        )
        return

    try:
        if CSV_FILE_PATH.suffix.lower() in {".xlsx", ".xls"}:
            sheet_name = os.getenv("AUDIT_SHEET_NAME")  # optional sheet override
            # Ensure openpyxl installed
            try:
                import openpyxl  # noqa: F401
            except ImportError:
                logger.error(
                    "❌ openpyxl not installed. Install dependency first (pip install openpyxl) or add to requirements.txt"
                )
                return
            read_kwargs = {"engine": "openpyxl"}
            if sheet_name:
                read_kwargs["sheet_name"] = sheet_name
            try:
                df = pd.read_excel(CSV_FILE_PATH, **read_kwargs)
            except Exception as e_primary:
                logger.warning(f"⚠️ pandas.read_excel failed ({e_primary}); attempting manual workbook parse fallback.")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(CSV_FILE_PATH, data_only=True)
                    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
                    rows = list(ws.values)
                    if not rows:
                        raise ValueError("Worksheet appears empty")
                    header = [str(c).strip() for c in rows[0]]
                    data_rows = rows[1:]
                    df = pd.DataFrame(data_rows, columns=header)
                except Exception as e_fallback:
                    logger.error(f"❌ Fallback workbook parse failed: {e_fallback}")
                    return
        else:
            df = pd.read_csv(CSV_FILE_PATH)
        df.columns = [c.strip() for c in df.columns]
    except Exception as e:
        logger.error(f"❌ Failed to read input file: {e}")
        return

    # 3. Filter Data
    # Ensure account ID comparison is string-to-string
    df['bill_account'] = df['bill_account'].astype(str)
    df_filtered = df[df['bill_account'] == str(TARGET_ACCOUNT)].copy()

    if df_filtered.empty:
        logger.warning(f"⚠️ No records found for account {TARGET_ACCOUNT}")
        return

    logger.info(f"📊 Found {len(df_filtered)} bills to audit.")

    # 4. Run Audit Loop
    audit_results = []
    
    for index, row in df_filtered.iterrows():
        # --- FORCE LOGIC OVERRIDE ---
        # In production, you might read 'service_class' from the DB. 
        # Here we force it as requested.
        row['service_class'] = FORCED_SC_CODE
        
        # Execute Math
        result = engine.calculate_expected_bill(row)
        
        # Build Result Row
        audit_row = row.to_dict()
        audit_row.update({
            "Audit_SC": result.get('sc_code'),
            "Expected_Bill": result.get('expected_bill'),
            "Variance": result.get('variance'),
            "Match_Status": "✅ Match" if abs(result.get('variance', 0)) < 0.10 else "❌ Mismatch",
            "Logic_Trace": " | ".join(result.get('trace', []))
        })
        
        audit_results.append(audit_row)

    # 5. Export Report
    df_final = pd.DataFrame(audit_results)
    
    # Organize columns for readability
    preferred_order = [
        'bill_account', 'bill_date', 'days_used', 'billed_kwh', 'billed_demand',
        'bill_amount', 'Expected_Bill', 'Variance', 'Match_Status', 'Logic_Trace'
    ]
    # Keep remaining columns at the end
    other_cols = [c for c in df_final.columns if c not in preferred_order]
    df_final = df_final[preferred_order + other_cols]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = OUTPUT_DIR / f"Audit_Result_{TARGET_ACCOUNT}_{timestamp}.xlsx"
    
    try:
        df_final.to_excel(output_filename, index=False)
        logger.info(f"✅ Success! Report saved to: {output_filename}")
        
        # Show a quick preview in the console
        print("\n--- Audit Preview ---")
        print(df_final[['bill_date', 'bill_amount', 'Expected_Bill', 'Variance', 'Match_Status']].head())
    except Exception as e:
        logger.error(f"❌ Failed to save Excel: {e}")

if __name__ == "__main__":
    run_audit_pipeline()